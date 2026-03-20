import os
from datetime import date, datetime
from io import BytesIO
import pandas as pd
from django.conf import settings
from django.db.models import Count, Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from aprendices.models import Aprendiz, Ficha, Inasistencia, AprendizResultado


class GeneradorReportes:
    """Genera reportes automáticos en Excel con formato profesional"""
    
    def __init__(self):
        self.hoy = date.today()
        
    def _aplicar_estilo_cabecera(self, ws, fila=1):
        """Aplica estilo a la fila de cabecera"""
        fill = PatternFill(start_color="00954a", end_color="00954a", fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=11)
        alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            bottom=Side(style='thin', color='000000')
        )
        
        for cell in ws[fila]:
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
            cell.border = border
    
    def _ajustar_columnas(self, ws):
        """Ajusta el ancho de las columnas automáticamente"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generar_reporte_inasistencias(self, ficha=None, fecha_desde=None, fecha_hasta=None):
        """
        Genera reporte consolidado de inasistencias
        
        Retorna: BytesIO con el archivo Excel
        """
        # Filtrar inasistencias
        queryset = Inasistencia.objects.select_related('aprendiz', 'ficha')
        
        if ficha:
            queryset = queryset.filter(ficha=ficha)
        if fecha_desde:
            queryset = queryset.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(fecha__lte=fecha_hasta)
        
        queryset = queryset.order_by('ficha__numero', 'fecha')
        
        # Crear DataFrame
        data = []
        for ina in queryset:
            data.append({
                'FICHA': ina.ficha.numero if ina.ficha else '',
                'INSTRUCTOR': ina.ficha.instructor if ina.ficha else '',
                'IDENTIFICACION APRENDIZ': ina.aprendiz.documento,
                'APRENDIZ': f"{ina.aprendiz.nombre} {ina.aprendiz.apellido}",
                'FECHA INICIO': ina.aprendiz.fecha_inicio.strftime('%d/%m/%Y') if ina.aprendiz.fecha_inicio else '',
                'FECHA F': ina.aprendiz.fecha_final.strftime('%d/%m/%Y') if ina.aprendiz.fecha_final else '',
                'CAN K': '',  # Campo vacío según tu imagen
                'JUSTIFICACION': ina.motivo or '',
            })
        
        df = pd.DataFrame(data)
        
        # Crear Excel con formato
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidado de Inasistencias"
        
        # Título
        ws.merge_cells('A1:H1')
        titulo_cell = ws['A1']
        titulo_cell.value = "Consolidado de Inasistencias - Aprendices por Ficha"
        titulo_cell.font = Font(bold=True, size=14, color="00954a")
        titulo_cell.alignment = Alignment(horizontal="center")
        
        # Subtítulo con fecha
        ws.merge_cells('A2:H2')
        subtitulo = ws['A2']
        subtitulo.value = f"Fecha del Reporte: {self.hoy.strftime('%d/%m/%Y')}"
        subtitulo.font = Font(size=10, italic=True)
        subtitulo.alignment = Alignment(horizontal="center")
        
        # Datos
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=4):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 4:  # Cabecera
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        self._ajustar_columnas(ws)
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def generar_reporte_juicios(self, ficha=None):
        """
        Genera reporte de juicios de evaluación
        
        Retorna: BytesIO con el archivo Excel
        """
        # Obtener datos
        queryset = AprendizResultado.objects.select_related(
            'aprendiz', 'aprendiz__ficha', 'resultado', 'resultado__competencia'
        )
        
        if ficha:
            queryset = queryset.filter(aprendiz__ficha=ficha)
        
        queryset = queryset.order_by('aprendiz__ficha__numero', 'aprendiz__documento')
        
        # Crear DataFrame
        data = []
        for ar in queryset:
            data.append({
                'Tipo': 'CC',
                'Documento': ar.aprendiz.documento,
                'Nombre': ar.aprendiz.nombre,
                'Apellidos': ar.aprendiz.apellido,
                'Estado': ar.aprendiz.estado_formacion,
                'Competencia': ar.resultado.competencia.codigo if ar.resultado.competencia else '',
                'Resultado de Aprendizaje': ar.resultado.codigo,
                'Juicio': ar.estado,
                'Fecha y Hora del Juicio': ar.fecha.strftime('%d/%m/%Y %H:%M') if ar.fecha else '',
                'Funcionario que registró': 'Sistema',
            })
        
        df = pd.DataFrame(data)
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Juicios"
        
        # Título
        ws.merge_cells('A1:J1')
        titulo = ws['A1']
        titulo.value = "Reporte de Juicios de Evaluación"
        titulo.font = Font(bold=True, size=14, color="00954a")
        titulo.alignment = Alignment(horizontal="center")
        
        # Info de la ficha si aplica
        if ficha:
            ws.merge_cells('A2:J2')
            info_ficha = ws['A2']
            info_ficha.value = f"Ficha de Caracterización: {ficha.numero}"
            info_ficha.font = Font(bold=True, size=11)
            info_ficha.alignment = Alignment(horizontal="left")
            
            ws.merge_cells('A3:J3')
            info_programa = ws['A3']
            info_programa.value = f"Denominación: {ficha.programa or 'Sin programa'}"
            info_programa.font = Font(size=10)
            
            inicio_datos = 5
        else:
            inicio_datos = 3
        
        # Datos
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=inicio_datos):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == inicio_datos:  # Cabecera
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        self._ajustar_columnas(ws)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def generar_reporte_circular120(self):
        """
        Genera reporte Circular 120 con casos vencidos y por certificar
        
        Retorna: BytesIO con el archivo Excel
        """
        hoy = self.hoy
        
        # Obtener aprendices relevantes
        por_certificar = Aprendiz.objects.filter(estado_formacion='POR_CERTIFICAR').select_related('ficha')
        
        productiva_vencida = Aprendiz.objects.filter(
            estado_formacion='ETAPA_PRODUCTIVA',
            fecha_fin_productiva__lt=hoy
        ).select_related('ficha')
        
        ficha_vencida = Aprendiz.objects.filter(
            ficha__fecha_fin__lt=hoy
        ).exclude(
            estado_formacion__in=['CERTIFICADO', 'CANCELADO']
        ).select_related('ficha')
        
        # Crear Excel con múltiples hojas
        wb = Workbook()
        
        # ===== HOJA 1: POR CERTIFICAR =====
        ws1 = wb.active
        ws1.title = "Por Certificar"
        
        data1 = []
        for ap in por_certificar:
            data1.append({
                'Documento': ap.documento,
                'Nombre Completo': f"{ap.nombre} {ap.apellido}",
                'Ficha': ap.ficha.numero if ap.ficha else '',
                'Programa': ap.ficha.programa if ap.ficha else '',
                'Estado': ap.estado_formacion,
                'Fecha Fin Productiva': ap.fecha_fin_productiva.strftime('%d/%m/%Y') if ap.fecha_fin_productiva else '',
                'Observaciones': ap.observaciones or '',
            })
        
        df1 = pd.DataFrame(data1)
        
        ws1['A1'] = "APRENDICES POR CERTIFICAR"
        ws1['A1'].font = Font(bold=True, size=14, color="00954a")
        ws1.merge_cells('A1:G1')
        
        for r_idx, row in enumerate(dataframe_to_rows(df1, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws1.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.fill = PatternFill(start_color="00954a", end_color="00954a", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
        
        self._ajustar_columnas(ws1)
        
        # ===== HOJA 2: PRODUCTIVA VENCIDA =====
        ws2 = wb.create_sheet("Productiva Vencida")
        
        data2 = []
        for ap in productiva_vencida:
            dias_venc = ap.dias_vencido()
            data2.append({
                'Documento': ap.documento,
                'Nombre Completo': f"{ap.nombre} {ap.apellido}",
                'Ficha': ap.ficha.numero if ap.ficha else '',
                'Programa': ap.ficha.programa if ap.ficha else '',
                'Fecha Fin Productiva': ap.fecha_fin_productiva.strftime('%d/%m/%Y') if ap.fecha_fin_productiva else '',
                'Días Vencido': dias_venc,
                'Nivel Urgencia': 'CRÍTICO' if dias_venc > 60 else 'MODERADO' if dias_venc > 30 else 'RECIENTE',
            })
        
        df2 = pd.DataFrame(data2)
        
        ws2['A1'] = "ETAPA PRODUCTIVA VENCIDA"
        ws2['A1'].font = Font(bold=True, size=14, color="d32f2f")
        ws2.merge_cells('A1:G1')
        
        for r_idx, row in enumerate(dataframe_to_rows(df2, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.fill = PatternFill(start_color="d32f2f", end_color="d32f2f", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
        
        self._ajustar_columnas(ws2)
        
        # ===== HOJA 3: FICHA VENCIDA =====
        ws3 = wb.create_sheet("Ficha Vencida")
        
        data3 = []
        for ap in ficha_vencida:
            dias_venc = ap.dias_vencido()
            data3.append({
                'Documento': ap.documento,
                'Nombre Completo': f"{ap.nombre} {ap.apellido}",
                'Ficha': ap.ficha.numero if ap.ficha else '',
                'Programa': ap.ficha.programa if ap.ficha else '',
                'Fecha Fin Ficha': ap.ficha.fecha_fin.strftime('%d/%m/%Y') if ap.ficha and ap.ficha.fecha_fin else '',
                'Días Vencido': dias_venc,
                'Estado': ap.estado_formacion,
            })
        
        df3 = pd.DataFrame(data3)
        
        ws3['A1'] = "FICHAS VENCIDAS SIN CERTIFICAR"
        ws3['A1'].font = Font(bold=True, size=14, color="f57c00")
        ws3.merge_cells('A1:G1')
        
        for r_idx, row in enumerate(dataframe_to_rows(df3, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws3.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.fill = PatternFill(start_color="f57c00", end_color="f57c00", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
        
        self._ajustar_columnas(ws3)
        
        # Guardar
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output


# Función helper para generar todos los reportes
def generar_todos_reportes():
    """
    Genera todos los reportes automáticamente y los guarda
    Retorna: dict con rutas de archivos generados
    """
    generador = GeneradorReportes()
    reportes_generados = {}
    
    # Directorio de reportes
    reportes_dir = os.path.join(settings.MEDIA_ROOT, 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # 1. Reporte de Inasistencias
    try:
        excel_inas = generador.generar_reporte_inasistencias()
        ruta_inas = os.path.join(reportes_dir, f'inasistencias_{timestamp}.xlsx')
        with open(ruta_inas, 'wb') as f:
            f.write(excel_inas.read())
        reportes_generados['inasistencias'] = ruta_inas
    except Exception as e:
        reportes_generados['inasistencias_error'] = str(e)
    
    # 2. Reporte de Juicios
    try:
        excel_juicios = generador.generar_reporte_juicios()
        ruta_juicios = os.path.join(reportes_dir, f'juicios_{timestamp}.xlsx')
        with open(ruta_juicios, 'wb') as f:
            f.write(excel_juicios.read())
        reportes_generados['juicios'] = ruta_juicios
    except Exception as e:
        reportes_generados['juicios_error'] = str(e)
    
    # 3. Reporte Circular 120
    try:
        excel_circular = generador.generar_reporte_circular120()
        ruta_circular = os.path.join(reportes_dir, f'circular120_{timestamp}.xlsx')
        with open(ruta_circular, 'wb') as f:
            f.write(excel_circular.read())
        reportes_generados['circular120'] = ruta_circular
    except Exception as e:
        reportes_generados['circular120_error'] = str(e)
    
    return reportes_generados